import openpyxl
import requests
import datetime
import asyncio
from openpyxl.writer.excel import save_workbook
from domains import domains

# Необходимые переменные
arr_url = domains
current_date = datetime.date.today().isoformat()
file_name = "data.xlsx"
arr_data = []

# Создаем книгу и листы excele
try:
    wb = openpyxl.load_workbook(file_name)
except:
    wb = openpyxl.Workbook()
ws = wb.create_sheet(f"{current_date}")

arr = []
# Функцияя получения ответов и занесения их в книгу
async def get_answer(site):
    try:
        res = requests.get(site)
        if str(res) != "<Response [200]>":
            print(site, res)
            arr.append([site, str(res)])
    except:
        print(site, "Ошибка")
        arr.append([site, "Ошибка"])


start_time = datetime.datetime.now()
count = len(domains)
row = 1
for url in arr_url:
    asyncio.run(get_answer(url))
    count -= 1
    print(count)
for date in arr:
    ws.cell(row=row, column=1).value = date[0]
    ws.cell(row=row, column=2).value = date[1]
    row += 1
# Сохраняем книгу
save_workbook(wb, file_name)
# Получаем время сбора данных
print(datetime.datetime.now() - start_time)
# Рабочая версия 1.1